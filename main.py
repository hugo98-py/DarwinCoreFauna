# -*- coding: utf-8 -*-
"""
FastAPI · Exportador Excel DwC-SMA (Render-ready)
------------------------------------------------
• Rellena la plantilla oficial DwC-SMA (.xlsx) con datos de Firestore.
• Mantiene 100 % el formato (openpyxl+lxml, edición celda-por-celda).
• Endpoint único:   /export?campana_id=…   →  {"download_url": "..."}
• Descargas servidas desde  /downloads/<archivo.xlsx>

Claves de producción:
  ▸ FIREBASE_KEY_B64   – variable de entorno con el JSON de la service-account
                         codificado en base-64 (una sola línea).
  ▸ Plantilla .xlsx     – debe residir junto a este archivo en el repo.
  ▸ Render Free plan    – ficheros de salida se guardan en /tmp.

Nuevo: middleware CORS para que peticiones desde navegadores no fallen.
"""

import os, re, base64, json, uuid, warnings
from pathlib import Path
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

from fastapi import FastAPI, Query, Request, HTTPException
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from openpyxl import load_workbook

import firebase_admin
from firebase_admin import credentials, firestore

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ────────────────────────────────  Rutas & constantes
ROOT_DIR      = Path(__file__).parent
TEMPLATE_PATH = ROOT_DIR / "FormatoBiodiversidadMonitoreoYLineaBase_v5.2.xlsx"
DOWNLOAD_DIR  = Path("/tmp/downloads")
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ────────────────────────────────  Firebase Init
B64 = os.environ.get("FIREBASE_KEY_B64")
if not B64:
    raise RuntimeError("FIREBASE_KEY_B64 env var is required")

cred_info = json.loads(base64.b64decode(B64))
if not firebase_admin._apps:
    firebase_admin.initialize_app(credentials.Certificate(cred_info))
db = firestore.client()

# ────────────────────────────────  FastAPI + CORS
app = FastAPI(title="Exporter DwC-SMA")

# ➜ MIDDLEWARE CORS (abierto a cualquier origen; ajusta si quieres)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],      # poner tu dominio en producción
    allow_methods=["GET"],
    allow_headers=["*"],
)

app.mount("/downloads", StaticFiles(directory=str(DOWNLOAD_DIR)), name="downloads")

# ────────────────────────────────  Utilidades
def _safe_filename(s: str) -> str:
    return re.sub(r"[^\w\-]+", "-", s)

LOCAL_TZ = ZoneInfo("America/Santiago")

def clean_dt(x):
    if isinstance(x, datetime):
        if x.tzinfo is not None:
            x = x.astimezone(LOCAL_TZ)   # conserva la hora real local
        return x.replace(tzinfo=None)
    return x

def fetch_df(collection: str, campana_id: str) -> pd.DataFrame:
    campana_id = campana_id.strip('"')
    ref        = db.collection(collection)
    docs       = list(ref.limit(1).stream())
    if not docs:
        return pd.DataFrame()

    # Si el documento tiene campanaID, filtramos
    if "campanaID" in docs[0].to_dict():
        ref = ref.where("campanaID", "==", campana_id)

    data = [{**d.to_dict(), "id": d.id} for d in ref.stream()]
    return pd.DataFrame(data).map(clean_dt)

def ymd(dt: Any):
    if pd.isna(dt):
        return None, None, None
    try:
        return int(dt.year), int(dt.month), int(dt.day)
    except Exception:
        return None, None, None

# ───────── Helpers de saneo numérico (soporta comas decimales) ─────────
def _to_none(val):
    """Convierte 999999 (int/float) y 'NO DATA' (str) en None."""
    if isinstance(val, (int, float)) and val == 999999:
        return None
    if isinstance(val, str) and val.strip().upper() == "NO DATA":
        return None
    return val

def _to_num(v):
    """Convierte strings con coma decimal/miles a float; valores raros → NaN."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    if isinstance(v, (int, float, np.number)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in {"nan", "none", "null"}:
        return np.nan
    # Si tiene comas y no puntos, interpretamos coma como decimal
    if s.count(",") >= 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    # Caso 1.234.567,89 → elimina separadores de miles
    if s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(s)
    except Exception:
        return np.nan

def _coerce_numeric_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            df[c] = df[c].map(_to_num)
    return df

# ────────────────────────────────  Generar Excel
def generar_excel(df_camp, df_met, df_reg, out_name: str) -> Path:
    wb   = load_workbook(TEMPLATE_PATH)
    ws_c = wb["Campaña"]
    ws_e = wb["EstacionReplica"]
    ws_o = wb["Ocurrencia"]

    # 1. Campaña -------------------------------------------------
    camp = df_camp.iloc[0].copy()
    camp["startDateCamp"] = pd.to_datetime(camp["startDateCamp"], errors="coerce")
    camp["endDateCamp"]   = pd.to_datetime(camp["endDateCamp"],   errors="coerce")
    y_i,m_i,d_i = ymd(camp["startDateCamp"])
    y_t,m_t,d_t = ymd(camp["endDateCamp"])

    for col,val in enumerate(
        [1, camp.get("Name"), camp.get("ncampana"), y_i,m_i,d_i, y_t,m_t,d_t], 1
    ):
        ws_c.cell(row=3, column=col, value=val)

    # 2. EstacionReplica (datos desde fila 2) --------------------
    df_met = df_met.copy()

    # Asegura columnas necesarias
    for col in ["startCoordTL", "endCoordTL", "centralCoordinate", "Type"]:
        if col not in df_met.columns:
            df_met[col] = None

    # Función simple para extraer lat/lon
    def get_lat(p): return getattr(p, "latitude", None) if pd.notna(p) else None
    def get_lon(p): return getattr(p, "longitude", None) if pd.notna(p) else None

    # Transecto Lineal → inicio y término
    mask_tl = df_met["Type"] == "Transecto Lineal"
    df_met.loc[mask_tl, "Latitud decimal inicio"]   = df_met["startCoordTL"].map(get_lat)
    df_met.loc[mask_tl, "Longitud decimal inicio"]  = df_met["startCoordTL"].map(get_lon)
    df_met.loc[mask_tl, "Latitud decimal término"]  = df_met["endCoordTL"].map(get_lat)
    df_met.loc[mask_tl, "Longitud decimal término"] = df_met["endCoordTL"].map(get_lon)

    # Otras metodologías → coordenada central
    mask_otras = ~mask_tl
    df_met.loc[mask_otras, "Latitud decimal central"]  = df_met["centralCoordinate"].map(get_lat)
    df_met.loc[mask_otras, "Longitud decimal central"] = df_met["centralCoordinate"].map(get_lon)

    # -------- LIMPIEZA NUMÉRICA Y SUPERFICIE (💥 fix del 500) --------
    # Coerciona columnas numéricas relevantes (ajusta si usas otras)
    df_met = _coerce_numeric_cols(
        df_met,
        ["Radio", "Ancho", "Largo", "Ancho M", "Largo M"]
    )

    # Superficie solo para 'Punto de Muestreo' con Radio válido
    mask_pm  = (df_met["Type"] == "Punto de Muestreo")
    radios   = pd.to_numeric(df_met["Radio"], errors="coerce")
    area_pm  = (np.pi * radios.pow(2)).round(0)
    df_met.loc[mask_pm, "Superficie (m2)"] = area_pm

    # ---------- B. Resto de transformaciones ----------
    df_met["Número Réplica"]     = df_met.groupby(["nameest", "Type"]).cumcount() + 1
    df_met["ID EstacionReplica"] = np.arange(1, len(df_met) + 1, dtype=int)
    df_met["Ecosistema nivel 2"] = df_met["ambienteest"]

    def build_tipo_mon(row):
        if row["Type"] in ("Transecto Lineal", "Play Back"):
            return f"{row['Type']} - {row.get('Clase')}"
        else:
            return row["Type"]

    df_met["Tipo de monitoreo"] = df_met.apply(build_tipo_mon, axis=1)

    # Renombrar a los títulos que exige la plantilla
    df_met = df_met.rename(columns={
        "nameest":       "Nombre estación",
        "Observaciones": "Descripción EstacionReplica",
        "Ancho":         "Ancho (m)",
        "Radio":         "Radio (m)",
        "region":        "Región",
        "provincia":     "Provincia",
        "comuna":        "Comuna",
        "localidad":     "Localidad",
    })

    # Diccionario “columna plantilla → índice Excel”
    cols_e = {
        "ID EstacionReplica": 1,  "Nombre estación": 2,     "Tipo de monitoreo": 3,
        "Número Réplica": 4,      "Descripción EstacionReplica": 5,
        "Largo (m)": 6,           "Ancho (m)": 7,           "Radio (m)": 8,
        "Superficie (m2)": 9,     "Latitud decimal central": 10,
        "Longitud decimal central": 11,
        "Latitud decimal inicio": 12,  "Longitud decimal inicio": 13,
        "Latitud decimal término": 14, "Longitud decimal término": 15,
        "Región": 16, "Provincia": 17, "Comuna": 18, "Localidad": 19,
        "Ecosistema nivel 2": 21,
    }

    # Nos aseguramos de que todas las columnas existan
    for c in cols_e:
        if c not in df_met.columns:
            df_met[c] = np.nan

    # Limpiamos filas previas en la hoja Excel (si las hubiera)
    if ws_e.max_row > 1:
        ws_e.delete_rows(2, ws_e.max_row - 1)

    # Escribimos fila por fila (índice limpio para no saltar filas)
    for excel_row, (_, fila) in enumerate(df_met.reset_index(drop=True).iterrows(), start=2):
        for col_name, col_idx in cols_e.items():
            ws_e.cell(row=excel_row, column=col_idx, value=fila[col_name])

    # 3. Ocurrencia ---------------------------------------------
    df_reg = df_reg.copy()
    df_reg["Time"] = pd.to_datetime(df_reg["Time"], errors="coerce")

    id_map = dict(zip(df_met.get("metodologiaID", []), df_met["ID EstacionReplica"]))
    if "metodologiaID" in df_reg.columns:
        df_reg["ID EstacionReplica"] = df_reg["metodologiaID"].map(id_map)
    else:
        df_reg["ID EstacionReplica"] = np.nan

    # Excluir tipos de la hoja Ocurrencia
    EXCLUDE_TYPES = {
        "Detección de Eco Localizaciones",   # ← nombre correcto
        "Trampas Sherman",
        "Trampas Cámara",
    }

    # Mapeamos metodologiaID ➜ Type (si existe)
    if "metodologiaID" in df_met.columns and "metodologiaID" in df_reg.columns:
        tipo_map = dict(zip(df_met["metodologiaID"], df_met["Type"]))
        df_reg["Type"] = df_reg["metodologiaID"].map(tipo_map)
        df_reg = df_reg[~df_reg["Type"].isin(EXCLUDE_TYPES)].copy()
        df_reg.drop(columns=["Type"], inplace=True, errors="ignore")
        df_reg.reset_index(drop=True, inplace=True)

    df_reg["Año del evento"]             = df_reg["Time"].dt.year
    df_reg["Mes del evento"]             = df_reg["Time"].dt.month
    df_reg["Día del evento"]             = df_reg["Time"].dt.day
    df_reg["Hora inicio evento (hh:mm)"] = df_reg["Time"].dt.strftime("%H:%M")
    df_reg["Hora registro"]              = df_reg["Hora inicio evento (hh:mm)"]
    df_reg["Latitud decimal registro"]   = df_reg["Coordinates"].apply(lambda c: getattr(c,"latitude",None))
    df_reg["Longitud decimal registro"]  = df_reg["Coordinates"].apply(lambda c: getattr(c,"longitude",None))

    df_reg["ID Campaña"]         = 1
    df_reg["Nombre campaña"]     = camp.get("Name")
    df_reg["Muestreado por"]     = "AMS Consultores"
    df_reg["Identificado por"]   = "AMS Consultores"
    df_reg["Epíteto específico"] = df_reg.get("epiteto")

    campos_o = {
        1:"ID Campaña",2:"Nombre campaña",3:"ID EstacionReplica",
        5:"Año del evento",6:"Mes del evento",7:"Día del evento",
        8:"Hora inicio evento (hh:mm)",
        9:"protocoloMuestreo",
        10:"tamanoMuestra",11:"unidadTamanoMuestra",
        14:"comentario",
        15:"reino",16:"division",17:"clase",18:"orden",
        19:"familia",20:"genero",
        22:"epiteto",24:"nombreComun",
        26:"estadoOrganismo",
        28:"parametro",29:"tipoCuantificacion",
        30:"valor",31:"unidadValor",
        32:"Latitud decimal registro",33:"Longitud decimal registro",
        34:"Hora registro",
        35:"condicionReproductiva",
        36:"sexo",37:"etapaVida",
        41:"tipoRegistro",
        44:"Muestreado por",45:"Identificado por",
    }
    for col in campos_o.values():
        if col not in df_reg.columns:
            df_reg[col] = ""

    if ws_o.max_row > 2:
        ws_o.delete_rows(3, ws_o.max_row - 2)

    for excel_row, (_, row) in enumerate(df_reg.reset_index(drop=True).iterrows(), start=3):
        for idx, col in campos_o.items():
            ws_o.cell(row=excel_row, column=idx, value=row[col])

    # Guardar archivo
    out_path = DOWNLOAD_DIR / out_name
    wb.save(out_path)
    return out_path

# ────────────────────────────────  Endpoint
@app.get("/export")
def export_excel(request: Request, campana_id: str = Query(...)):
    df_camp = fetch_df("campana",     campana_id)
    df_met  = fetch_df("Metodologia", campana_id)
    df_reg  = fetch_df("Registro",    campana_id)

    if df_camp.empty or df_met.empty or df_reg.empty:
        raise HTTPException(status_code=404, detail="No hay datos para la campaña.")

    # Bloque limpiador (sin applymap deprecado)
    df_met = df_met.apply(lambda col: col.map(_to_none))
    df_reg = df_reg.apply(lambda col: col.map(_to_none))

    filename = f"DWC_{_safe_filename(campana_id)}_{uuid.uuid4().hex[:6]}.xlsx"
    path     = generar_excel(df_camp, df_met, df_reg, filename)

    download_url = f"{str(request.base_url).rstrip('/')}/downloads/{path.name}"
    return JSONResponse({"download_url": download_url})











